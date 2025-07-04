"""
Enhanced File Processor for WiFi Data Automation
Handles CSV to Excel conversion with proper header mapping and Excel 2007 format
"""

import pandas as pd
import csv
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
import os
import shutil
import re

# Local imports
from config.settings import FILE_CONFIG, WIFI_CONFIG, SCHEDULE_CONFIG
from core.logger import logger


class WiFiDataProcessor:
    """
    Enhanced WiFi Data Processor for CSV to Excel conversion
    """
    
    def __init__(self, execution_id: Optional[str] = None):
        self.execution_id = execution_id or f"processor_{int(datetime.now().timestamp())}"
        self.processed_files = []
        self.processing_stats = {
            'total_files': 0,
            'successful_conversions': 0,
            'failed_conversions': 0,
            'total_records': 0,
            'merged_records': 0
        }
        
        # Initialize directories
        self.current_date = datetime.now()
        self.date_folder = self.current_date.strftime(FILE_CONFIG['date_folder_format'])
        
        self.csv_dir = Path(FILE_CONFIG['csv_storage_path'].format(date_folder=self.date_folder))
        self.excel_dir = Path(FILE_CONFIG['excel_storage_path'].format(date_folder=self.date_folder))
        
        # Create directories
        self.csv_dir.mkdir(parents=True, exist_ok=True)
        self.excel_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"WiFi Data Processor initialized: {self.execution_id}", "FileProcessor", self.execution_id)
        logger.info(f"CSV Directory: {self.csv_dir}", "FileProcessor", self.execution_id)
        logger.info(f"Excel Directory: {self.excel_dir}", "FileProcessor", self.execution_id)
    
    def process_csv_file(self, csv_file_path: str) -> Dict[str, Any]:
        """
        Process a single CSV file and convert to standardized format
        """
        try:
            csv_path = Path(csv_file_path)
            logger.info(f"Processing CSV file: {csv_path.name}", "FileProcessor", self.execution_id)
            
            # Read CSV file
            df = pd.read_csv(csv_path, encoding='utf-8')
            
            if df.empty:
                logger.warning(f"CSV file is empty: {csv_path.name}", "FileProcessor", self.execution_id)
                return {
                    'success': False,
                    'error': 'Empty CSV file',
                    'file_path': str(csv_path),
                    'record_count': 0
                }
            
            # Log original columns
            logger.info(f"Original columns: {list(df.columns)}", "FileProcessor", self.execution_id)
            
            # Clean column names (remove extra spaces, standardize)
            df.columns = df.columns.str.strip()
            
            # Map headers according to configuration
            mapped_df = self._map_csv_headers(df)
            
            if mapped_df is None:
                return {
                    'success': False,
                    'error': 'Header mapping failed',
                    'file_path': str(csv_path),
                    'record_count': len(df)
                }
            
            # Clean and validate data
            cleaned_df = self._clean_data(mapped_df)
            
            # Add metadata
            cleaned_df['Source_File'] = csv_path.name
            cleaned_df['Processing_Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            result = {
                'success': True,
                'file_path': str(csv_path),
                'record_count': len(cleaned_df),
                'processed_data': cleaned_df,
                'original_columns': list(df.columns),
                'mapped_columns': list(cleaned_df.columns)
            }
            
            logger.success(f"Successfully processed {csv_path.name}: {len(cleaned_df)} records", "FileProcessor", self.execution_id)
            return result
            
        except Exception as e:
            logger.error(f"Failed to process CSV file {csv_file_path}: {e}", "FileProcessor", self.execution_id)
            return {
                'success': False,
                'error': str(e),
                'file_path': csv_file_path,
                'record_count': 0
            }
    
    def _map_csv_headers(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """
        Map CSV headers to Excel format according to configuration
        """
        try:
            logger.info("Mapping CSV headers", "FileProcessor", self.execution_id)
            
            # Get header mapping configuration
            header_mapping = FILE_CONFIG['header_mapping']
            source_headers = FILE_CONFIG['source_csv_headers']
            
            # Create a flexible mapping that handles variations
            flexible_mapping = {}
            
            for source_header, target_header in header_mapping.items():
                # Find matching column in DataFrame (case-insensitive, partial match)
                for col in df.columns:
                    col_clean = col.strip().lower()
                    source_clean = source_header.lower()
                    
                    # Exact match
                    if col_clean == source_clean:
                        flexible_mapping[col] = target_header
                        break
                    
                    # Partial match for common variations
                    elif 'hostname' in col_clean and 'hostname' in source_clean:
                        flexible_mapping[col] = target_header
                        break
                    elif 'ip' in col_clean and 'ip address' in source_clean:
                        flexible_mapping[col] = target_header
                        break
                    elif 'mac' in col_clean and 'mac address' in source_clean:
                        flexible_mapping[col] = target_header
                        break
                    elif 'ssid' in col_clean or 'wlan' in col_clean and 'wlan' in source_clean:
                        flexible_mapping[col] = target_header
                        break
                    elif 'ap' in col_clean and 'ap mac' in source_clean:
                        flexible_mapping[col] = target_header
                        break
                    elif 'up' in col_clean and 'data rate (up)' in source_clean:
                        flexible_mapping[col] = target_header
                        break
                    elif 'down' in col_clean and 'data rate (down)' in source_clean:
                        flexible_mapping[col] = target_header
                        break
            
            logger.info(f"Header mapping: {flexible_mapping}", "FileProcessor", self.execution_id)
            
            if not flexible_mapping:
                logger.warning("No header mappings found, using original columns", "FileProcessor", self.execution_id)
                return df
            
            # Rename columns
            mapped_df = df.rename(columns=flexible_mapping)
            
            # Ensure all required columns exist
            excel_headers = FILE_CONFIG['excel_headers']
            for header in excel_headers:
                if header not in mapped_df.columns:
                    mapped_df[header] = ''  # Add missing columns with empty values
            
            # Reorder columns according to Excel header order
            available_columns = [col for col in excel_headers if col in mapped_df.columns]
            other_columns = [col for col in mapped_df.columns if col not in excel_headers]
            
            final_columns = available_columns + other_columns
            mapped_df = mapped_df[final_columns]
            
            logger.success(f"Headers mapped successfully: {list(mapped_df.columns)}", "FileProcessor", self.execution_id)
            return mapped_df
            
        except Exception as e:
            logger.error(f"Header mapping failed: {e}", "FileProcessor", self.execution_id)
            return None
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and validate data
        """
        try:
            logger.info("Cleaning data", "FileProcessor", self.execution_id)
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            # Clean string columns
            string_columns = df.select_dtypes(include=['object']).columns
            for col in string_columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')
                df[col] = df[col].replace('None', '')
            
            # Validate MAC addresses
            if 'MAC_Address' in df.columns:
                df['MAC_Address'] = df['MAC_Address'].apply(self._validate_mac_address)
            
            if 'AP_MAC' in df.columns:
                df['AP_MAC'] = df['AP_MAC'].apply(self._validate_mac_address)
            
            # Validate IP addresses
            if 'IP_Address' in df.columns:
                df['IP_Address'] = df['IP_Address'].apply(self._validate_ip_address)
            
            # Clean data rates
            if 'Upload' in df.columns:
                df['Upload'] = df['Upload'].apply(self._clean_data_rate)
            
            if 'Download' in df.columns:
                df['Download'] = df['Download'].apply(self._clean_data_rate)
            
            logger.success(f"Data cleaned successfully: {len(df)} records", "FileProcessor", self.execution_id)
            return df
            
        except Exception as e:
            logger.error(f"Data cleaning failed: {e}", "FileProcessor", self.execution_id)
            return df
    
    def _validate_mac_address(self, mac: str) -> str:
        """
        Validate and format MAC address
        """
        if not mac or mac == '' or mac == 'nan':
            return ''
        
        # Remove common separators and spaces
        mac_clean = re.sub(r'[:-\s]', '', str(mac))
        
        # Check if it's a valid MAC address format
        if len(mac_clean) == 12 and all(c in '0123456789abcdefABCDEF' for c in mac_clean):
            # Format as XX:XX:XX:XX:XX:XX
            return ':'.join(mac_clean[i:i+2] for i in range(0, 12, 2)).upper()
        
        return str(mac)  # Return original if not valid
    
    def _validate_ip_address(self, ip: str) -> str:
        """
        Validate IP address
        """
        if not ip or ip == '' or ip == 'nan':
            return ''
        
        # Basic IP validation
        parts = str(ip).split('.')
        if len(parts) == 4:
            try:
                return '.'.join(str(int(part)) for part in parts if 0 <= int(part) <= 255)
            except:
                pass
        
        return str(ip)  # Return original if not valid
    
    def _clean_data_rate(self, rate: str) -> str:
        """
        Clean data rate values
        """
        if not rate or rate == '' or rate == 'nan':
            return '0'
        
        # Extract numeric value
        rate_str = str(rate).strip()
        numeric_match = re.search(r'(\d+(?:\.\d+)?)', rate_str)
        
        if numeric_match:
            return numeric_match.group(1)
        
        return '0'
    
    def merge_csv_files(self, csv_files: List[str], slot_name: str = None) -> Dict[str, Any]:
        """
        Merge multiple CSV files into a single dataset
        """
        try:
            slot_name = slot_name or f"slot_{int(datetime.now().timestamp())}"
            logger.info(f"Merging {len(csv_files)} CSV files for slot: {slot_name}", "FileProcessor", self.execution_id)
            
            all_dataframes = []
            processing_results = []
            
            for csv_file in csv_files:
                result = self.process_csv_file(csv_file)
                processing_results.append(result)
                
                if result['success']:
                    all_dataframes.append(result['processed_data'])
                    self.processing_stats['successful_conversions'] += 1
                else:
                    self.processing_stats['failed_conversions'] += 1
                
                self.processing_stats['total_files'] += 1
            
            if not all_dataframes:
                return {
                    'success': False,
                    'error': 'No valid CSV files to merge',
                    'slot_name': slot_name,
                    'processing_results': processing_results
                }
            
            # Merge all dataframes
            merged_df = pd.concat(all_dataframes, ignore_index=True)
            
            # Remove duplicates based on MAC address
            if 'MAC_Address' in merged_df.columns:
                initial_count = len(merged_df)
                merged_df = merged_df.drop_duplicates(subset=['MAC_Address'], keep='first')
                duplicates_removed = initial_count - len(merged_df)
                
                if duplicates_removed > 0:
                    logger.info(f"Removed {duplicates_removed} duplicate records", "FileProcessor", self.execution_id)
            
            # Sort by hostname
            if 'Hostname' in merged_df.columns:
                merged_df = merged_df.sort_values('Hostname')
            
            # Update statistics
            self.processing_stats['total_records'] = sum(r['record_count'] for r in processing_results if r['success'])
            self.processing_stats['merged_records'] = len(merged_df)
            
            result = {
                'success': True,
                'slot_name': slot_name,
                'merged_data': merged_df,
                'record_count': len(merged_df),
                'processing_results': processing_results,
                'statistics': self.processing_stats.copy()
            }
            
            logger.success(f"Successfully merged CSV files: {len(merged_df)} total records", "FileProcessor", self.execution_id)
            return result
            
        except Exception as e:
            logger.error(f"CSV merge failed: {e}", "FileProcessor", self.execution_id)
            return {
                'success': False,
                'error': str(e),
                'slot_name': slot_name,
                'processing_results': []
            }
    
    def create_excel_file(self, merged_data: pd.DataFrame, slot_name: str = None) -> Dict[str, Any]:
        """
        Create Excel file in 2007 format (.xlsx) with proper formatting
        """
        try:
            slot_name = slot_name or f"slot_{int(datetime.now().timestamp())}"
            
            # Generate filename
            date_str = datetime.now().strftime('%d%m%Y')
            filename = FILE_CONFIG['excel_filename_template'].format(date=date_str)
            excel_path = self.excel_dir / filename
            
            logger.info(f"Creating Excel file: {excel_path}", "FileProcessor", self.execution_id)
            
            # Create Excel file using openpyxl for better formatting
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "WiFi Users Data"
            
            # Get Excel headers in correct order
            excel_headers = FILE_CONFIG['excel_headers']
            
            # Filter and reorder merged data columns
            available_columns = [col for col in excel_headers if col in merged_data.columns]
            reordered_data = merged_data[available_columns]
            
            # Write headers
            for col_idx, header in enumerate(available_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for row_idx, (_, row) in enumerate(reordered_data.iterrows(), 2):
                for col_idx, header in enumerate(available_columns, 1):
                    value = row[header]
                    # Handle NaN values
                    if pd.isna(value):
                        value = ""
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # Save Excel file
            wb.save(excel_path)
            
            # Verify file was created and get size
            if excel_path.exists():
                file_size = excel_path.stat().st_size
                logger.success(f"Excel file created successfully: {excel_path.name} ({file_size} bytes)", "FileProcessor", self.execution_id)
                
                return {
                    'success': True,
                    'excel_path': str(excel_path),
                    'filename': filename,
                    'record_count': len(reordered_data),
                    'file_size': file_size,
                    'columns': available_columns,
                    'slot_name': slot_name
                }
            else:
                return {
                    'success': False,
                    'error': 'Excel file was not created',
                    'slot_name': slot_name
                }
                
        except Exception as e:
            logger.error(f"Excel file creation failed: {e}", "FileProcessor", self.execution_id)
            return {
                'success': False,
                'error': str(e),
                'slot_name': slot_name
            }
    
    def process_complete_workflow(self, csv_files: List[str], slot_name: str = None) -> Dict[str, Any]:
        """
        Execute complete CSV to Excel workflow
        """
        try:
            slot_name = slot_name or f"slot_{int(datetime.now().timestamp())}"
            logger.info(f"Starting complete workflow for slot: {slot_name}", "FileProcessor", self.execution_id)
            
            # Step 1: Merge CSV files
            merge_result = self.merge_csv_files(csv_files, slot_name)
            
            if not merge_result['success']:
                return {
                    'success': False,
                    'error': f"CSV merge failed: {merge_result.get('error', 'Unknown error')}",
                    'slot_name': slot_name,
                    'step_failed': 'merge'
                }
            
            # Step 2: Create Excel file
            excel_result = self.create_excel_file(merge_result['merged_data'], slot_name)
            
            if not excel_result['success']:
                return {
                    'success': False,
                    'error': f"Excel creation failed: {excel_result.get('error', 'Unknown error')}",
                    'slot_name': slot_name,
                    'step_failed': 'excel_creation',
                    'merge_result': merge_result
                }
            
            # Compile final results
            final_result = {
                'success': True,
                'slot_name': slot_name,
                'execution_id': self.execution_id,
                'timestamp': datetime.now().isoformat(),
                'excel_file': excel_result['excel_path'],
                'excel_filename': excel_result['filename'],
                'total_records': excel_result['record_count'],
                'csv_files_processed': len(csv_files),
                'successful_csv_files': merge_result['statistics']['successful_conversions'],
                'failed_csv_files': merge_result['statistics']['failed_conversions'],
                'columns': excel_result['columns'],
                'file_size': excel_result['file_size'],
                'merge_result': merge_result,
                'excel_result': excel_result
            }
            
            logger.success(f"Complete workflow finished successfully: {excel_result['filename']} with {excel_result['record_count']} records", 
                          "FileProcessor", self.execution_id)
            
            return final_result
            
        except Exception as e:
            logger.error(f"Complete workflow failed: {e}", "FileProcessor", self.execution_id)
            return {
                'success': False,
                'error': str(e),
                'slot_name': slot_name,
                'execution_id': self.execution_id,
                'timestamp': datetime.now().isoformat()
            }
    
    def cleanup_old_files(self, retention_days: int = 30) -> Dict[str, Any]:
        """
        Clean up old CSV and Excel files
        """
        try:
            logger.info(f"Cleaning up files older than {retention_days} days", "FileProcessor", self.execution_id)
            
            cutoff_date = datetime.now() - timedelta(days=retention_days)
            deleted_files = []
            
            # Clean CSV files
            for csv_file in self.csv_dir.rglob("*.csv"):
                if csv_file.stat().st_mtime < cutoff_date.timestamp():
                    try:
                        csv_file.unlink()
                        deleted_files.append(str(csv_file))
                    except Exception as e:
                        logger.warning(f"Failed to delete {csv_file}: {e}", "FileProcessor", self.execution_id)
            
            # Clean Excel files
            for excel_file in self.excel_dir.rglob("*.xlsx"):
                if excel_file.stat().st_mtime < cutoff_date.timestamp():
                    try:
                        excel_file.unlink()
                        deleted_files.append(str(excel_file))
                    except Exception as e:
                        logger.warning(f"Failed to delete {excel_file}: {e}", "FileProcessor", self.execution_id)
            
            logger.success(f"Cleanup completed: {len(deleted_files)} files deleted", "FileProcessor", self.execution_id)
            
            return {
                'success': True,
                'deleted_files': deleted_files,
                'deleted_count': len(deleted_files)
            }
            
        except Exception as e:
            logger.error(f"Cleanup failed: {e}", "FileProcessor", self.execution_id)
            return {
                'success': False,
                'error': str(e),
                'deleted_count': 0
            }


# Convenience functions for external use
def process_wifi_data(csv_files: List[str], slot_name: str = None) -> Dict[str, Any]:
    """
    Process WiFi data from CSV files to Excel format
    """
    processor = WiFiDataProcessor()
    return processor.process_complete_workflow(csv_files, slot_name)


def merge_slot_data(slot_data: Dict[str, List[str]]) -> Dict[str, Any]:
    """
    Merge data from multiple slots into a single Excel file
    """
    try:
        logger.info(f"Merging data from {len(slot_data)} slots", "FileProcessor", "merge_slots")
        
        all_csv_files = []
        for slot_name, csv_files in slot_data.items():
            all_csv_files.extend(csv_files)
        
        if not all_csv_files:
            return {
                'success': False,
                'error': 'No CSV files to merge'
            }
        
        # Process all files together
        processor = WiFiDataProcessor()
        result = processor.process_complete_workflow(all_csv_files, "merged_slots")
        
        if result['success']:
            logger.success(f"Successfully merged {len(slot_data)} slots into {result['excel_filename']}", 
                          "FileProcessor", "merge_slots")
        
        return result
        
    except Exception as e:
        logger.error(f"Slot data merge failed: {e}", "FileProcessor", "merge_slots")
        return {
            'success': False,
            'error': str(e)
        }